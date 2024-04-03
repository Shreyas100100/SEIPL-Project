from cProfile import label
import tkinter as tk
from tkinter import ttk, StringVar, messagebox, filedialog
import pandas as pd
import customtkinter
from openpyxl import load_workbook


def choose_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        with open("path.txt", "w") as f:
            f.write(file_path)
        label.config(text="Selected File: " + file_path)
        
def create_record():
    create_window = tk.Toplevel(root)
    create_window.title("Create Record")

    create_frame = customtkinter.CTkScrollableFrame(create_window, orientation="horizontal")
    create_frame.pack(fill=customtkinter.BOTH, expand=True)

    entry_fields = {}
    max_columns = 5  # Maximum number of columns per row
    num_columns = min(len(df.columns[:70]), max_columns)  # Display only the first 30 columns
    for index, col_name in enumerate(df.columns[:70]):
        row = index // max_columns
        col = index % max_columns
        label = customtkinter.CTkLabel(create_frame, text=col_name)
        label.grid(row=row, column=col * 2, padx=5, pady=5, sticky=tk.E)
        entry = customtkinter.CTkEntry(create_frame)
        entry.grid(row=row, column=col * 2 + 1, padx=5, pady=5, sticky=tk.W)
        entry_fields[col_name] = entry

    def save_new_record():
        new_values = [entry.get() for entry in entry_fields.values()]
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet.append(new_values)
        wb.save(filename=file_path)
        messagebox.showinfo("Record Created", "New record has been created successfully.")
        create_window.destroy()

    def import_file():
        new_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if new_file:
            df_new = pd.read_excel(new_file)
            d1 = pd.read_excel(file_path)
            # Append data to existing Excel file
            combined_df = pd.concat([d1, df_new], ignore_index=True)
            # Save the combined DataFrame to the same Excel file
            combined_df.to_excel(file_path, index=False)
            print('Done !')
            messagebox.showinfo("Import Status","File Imported successfully")
            refresh()
        else:
            messagebox.showinfo(" Error Fetching File")

        create_window.destroy()

    import_button = customtkinter.CTkButton(create_window, text="Import File", command=import_file)
    import_button.pack(pady=10)

    save_button = customtkinter.CTkButton(create_window, text="Save", command=save_new_record)
    save_button.pack(pady=10)
    
def show_filter_message(filtered_df):
    count = len(filtered_df)
    messagebox.showinfo("Filter Applied", f"Filter applied. {count} rows found.")


def filter_data():
    global df
    filtered_df = df.copy()
    for column, entry_var in zip(df.columns, filter_vars):
        value = entry_var.get()
        if value.strip():  # Check if the entry is not empty
            filtered_df = filtered_df[filtered_df[column].astype(str).str.contains(value, case=False)]
    update_treeview(filtered_df)
    show_filter_message(filtered_df)


def copy_data():
    selected_items = my_tree.selection()
    if selected_items:
        selected_rows = [my_tree.item(item, 'values') for item in selected_items]
        # Convert the selected rows to a tab-separated string for copying
        copied_text = '\n'.join(['\t'.join(map(str, row)) for row in selected_rows])
        root.clipboard_clear()
        root.clipboard_append(copied_text)
        root.update()
        messagebox.showinfo("Copied", "Selected rows have been copied to the clipboard.")
    else:
        messagebox.showinfo("No Selection", "Please select rows to copy.")

def toggle_theme():
    global is_dark_mode
    is_dark_mode = not is_dark_mode
    update_theme()

def update_theme():
    if is_dark_mode:
        root.config(bg="#1E1E1E")
        # Apply dark mode theme to customtkinter
        customtkinter.set_appearance_mode("dark")
        customtkinter.set_default_color_theme("green")
    else:
        root.config(bg="white")
        # Apply light mode theme to customtkinter
        customtkinter.set_appearance_mode("light")
        customtkinter.set_default_color_theme("green")

customtkinter.set_appearance_mode("light")
customtkinter.set_default_color_theme("green")
root = customtkinter.CTk()
root.title('CMDB')

with open("path.txt", "r") as f:
        file_path = f.read()
df = pd.read_excel(file_path)

def refresh():
    global df
    df = pd.read_excel(file_path)
    update_treeview(df)

def save_as_excel():
    # Ask user for file save location and name
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        # Get filtered data
        filtered_df = get_filtered_data()
        # Save filtered data to Excel file
        filtered_df.to_excel(file_path, index=False)
        messagebox.showinfo("Saved", "Filtered data has been saved successfully.")

def get_filtered_data():
    filtered_df = df.copy()
    for column, entry_var in zip(df.columns, filter_vars):
        value = entry_var.get()
        if value.strip():  # Check if the entry is not empty
            filtered_df = filtered_df[filtered_df[column].astype(str).str.contains(value, case=False)]
    return filtered_df

# Frame to contain filter buttons
filter_frame = customtkinter.CTkScrollableFrame(root, orientation="horizontal" ,height=600)
filter_frame.pack(side="top", fill="x", pady=20, padx=30)

frm = customtkinter.CTkFrame(root)
dfr=pd.read_excel(file_path)
cnt=len(dfr)
l1 = customtkinter.CTkLabel(root,text=cnt,)
l1.pack(side="left",padx=5, pady=1)

l2 = customtkinter.CTkLabel(root,text=" Rows Displayed")
l2.pack(side="left",padx=5, pady=1)

filter_vars = []
filter_entries = []
for idx, column in enumerate(df.columns[0:70]):  # Display only the first 30 columns
    filter_label = customtkinter.CTkLabel(filter_frame, text=column)
    filter_label.grid(row=1, column=idx+1, padx=5)
    
    filter_var = StringVar()
    filter_vars.append(filter_var)
    filter_entry = customtkinter.CTkEntry(filter_frame, textvariable=filter_var)
    filter_entry.grid(row=0, column=idx+1, padx=5)
    filter_entries.append(filter_entry)

# Buttons
apply_filters_button = customtkinter.CTkButton(root, text="RUN", command=filter_data)
apply_filters_button.pack(side="left", pady=5)

create_button = customtkinter.CTkButton(root, text="CREATE",command=create_record)
create_button.pack(side="left", padx=5)

update_button = customtkinter.CTkButton(root, text="UPDATE")
update_button.pack(side="left", padx=5)

copy_button = customtkinter.CTkButton(root, text="COPY",command=copy_data)
copy_button.pack(side="left", padx=5)

refresh_button = customtkinter.CTkButton(root, text="REFRESH",command=refresh)
refresh_button.pack(side="left", padx=5)

choose_button = customtkinter.CTkButton(root, text="Choose File", command=choose_file)
choose_button.pack(side="left",padx=5)

save_excel_button = customtkinter.CTkButton(root, text="Save as Excel", command=save_as_excel)
save_excel_button.pack(side="left", padx=5)

# TreeView
my_tree = ttk.Treeview(filter_frame, height=45, selectmode="extended", style="Treeview")
my_tree.grid(row=2, column=1, columnspan=70, pady=20)  # Displaying 30 columns

my_tree.delete(*my_tree.get_children())
my_tree['column'] = list(df.columns[:70])  # Displaying 30 columns
my_tree['show'] = 'headings'

# Show headings
for col in my_tree['column']:
    my_tree.heading(col, text=col)

# Show Data
def update_treeview(new_df):
    my_tree.delete(*my_tree.get_children())
    df_rows = new_df.iloc[:, :70].to_numpy().tolist()  # Displaying 30 columns
    for row in df_rows:
        my_tree.insert("", "end", values=row)

update_treeview(df)

# Function to handle selection of filter entry
def on_entry_select(event):
    entry = event.widget
    canvas = entry.master.master
    scrollbar = canvas.scrollbar
    bbox = canvas.bbox(tk.ALL)
    item_y = entry.winfo_rooty() - canvas.winfo_rooty() + entry.winfo_height()  # Bottom y-coordinate of the entry relative to the canvas
    view_position = scrollbar.get()[0]  # Get the current view position of the canvas
    
    if item_y > bbox[3] or item_y < bbox[1]:  # If the entry is out of view
        # Calculate the fraction of canvas to scroll
        fraction = (entry.winfo_rooty() - canvas.winfo_rooty() + entry.winfo_height()) / bbox[3]
        new_view_position = view_position + fraction
        scrollbar.set(new_view_position, new_view_position)

# Set tree style
style = ttk.Style()
style.configure("Treeview.Heading", font=("TkDefaultFont", 10), borderwidth=2)
style.configure("Treeview", font=("TkDefaultFont", 10), borderwidth=2)

# Function to apply alternating row colors
def apply_alternating_row_colors():
    for idx, item in enumerate(my_tree.get_children()):
        if idx % 2 == 0:
            my_tree.item(item, tags=('even_row',))
        else:
            my_tree.item(item, tags=('odd_row',))

    my_tree.tag_configure('even_row', background='#F0F0F0')  # Light grey
    my_tree.tag_configure('odd_row', background='white')

# Call the function to apply alternating row colors
apply_alternating_row_colors()


# Function to add vertical lines between columns
def add_vertical_lines():
    num_columns = len(df.columns[:70])  # Display only the first 30 columns
    for i in range(num_columns - 1):
        my_tree.heading(df.columns[i], text=df.columns[i] + ' |')  # Add separator after each column

# Call the function to add vertical lines
add_vertical_lines()

# Vertical Scrollbar
vsb = customtkinter.CTkScrollbar(filter_frame, command=my_tree.yview)
vsb.grid(row=2, column=0, sticky='ns')  # Displaying 30 columns
my_tree.configure(yscrollcommand=vsb.set)
# Sliding button for light and dark mode
is_dark_mode = True
theme_button = customtkinter.CTkButton(root, text="Dark Mode", command=toggle_theme)
theme_button.pack(padx=5, side="left")

# Function to handle selection
def on_select(event):
    item = event.widget.selection()[0]
    values = event.widget.item(item, 'values')
    print("Selected values:", values)

# Bind the selection event
my_tree.bind("<<TreeviewSelect>>", on_select)

# Function to hide column
def hide_column(event):
    column_id = my_tree.identify_column(event.x)
    column_index = int(column_id.replace("#", ""))
    column_name = my_tree['columns'][column_index]
    my_tree.column(column_name, width=0)  # Hide the column by setting its width to 0

# Bind the right-click event on column headers
for col in my_tree['column']:
    my_tree.heading(col, text=col, command=lambda: hide_column(col))
def create_record():
    create_window = tk.Toplevel(root)
    create_window.title("Create Record")

    create_frame = customtkinter.CTkScrollableFrame(create_window, orientation="horizontal")
    create_frame.pack(fill=customtkinter.BOTH, expand=True)

    entry_fields = {}
    max_columns = 5  # Maximum number of columns per row
    num_columns = min(len(df.columns[:70]), max_columns)  # Display only the first 30 columns
    for index, col_name in enumerate(df.columns[:70]):
        row = index // max_columns
        col = index % max_columns
        label = customtkinter.CTkLabel(create_frame, text=col_name)
        label.grid(row=row, column=col * 2, padx=5, pady=5, sticky=tk.E)
        entry = customtkinter.CTkEntry(create_frame)
        entry.grid(row=row, column=col * 2 + 1, padx=5, pady=5, sticky=tk.W)
        entry_fields[col_name] = entry

    def save_new_record():
        new_values = [entry.get() for entry in entry_fields.values()]
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        sheet.append(new_values)
        wb.save(filename=file_path)
        update_treeview(df)
        update_excel_file(df)
        refresh()
        messagebox.showinfo("Record Created", "New record has been created successfully.")
        refresh
        create_window.destroy()

    save_button = customtkinter.CTkButton(create_window, text="Save", command=save_new_record)
    save_button.pack(pady=10)
    
# Function to update record
def update_record():
    selected_items = my_tree.selection()
    if selected_items:
        # Open a new window for updating
        update_window = tk.Toplevel(root)
        update_window.title("Update Record")

        selected_row = my_tree.item(selected_items[0], 'values')

        # Create a frame inside the update window to contain entry fields
        update_frame = customtkinter.CTkScrollableFrame(update_window,orientation="horizontal",)
        update_frame.pack(fill=customtkinter.BOTH, expand=True)

        # Create entry fields for each column
        entry_fields = {}
        max_columns = 5  # Maximum number of columns per row
        num_columns = min(len(df.columns[:70]), max_columns)  # Display only the first 30 columns
        for index, col_name in enumerate(df.columns[:70]):
            row = index // max_columns
            col = index % max_columns
            label = customtkinter.CTkLabel(update_frame, text=col_name)
            label.grid(row=row, column=col * 2, padx=5, pady=5, sticky=tk.E)
            entry = customtkinter.CTkEntry(update_frame)
            entry.grid(row=row, column=col * 2 + 1, padx=5, pady=5, sticky=tk.W)
            entry.insert(0, selected_row[index])
            entry_fields[col_name] = entry

        # Create a button to save the updated record
        def save_updated_record():
            updated_values = [entry.get() for entry in entry_fields.values()]
            selected_item = my_tree.selection()[0]
            item_index = int(my_tree.index(selected_item))
            for col_name, value in zip(df.columns[:70], updated_values):
                df.at[item_index, col_name] = value  
            update_treeview(df)
            update_excel_file(df)  
            update_window.destroy()

        save_button = customtkinter.CTkButton(update_window, text="Save", command=save_updated_record)
        save_button.pack(pady=10)
    else:
        messagebox.showwarning("Update", "No row selected.")


update_button.configure(command=update_record)

def update_excel_file(dataframe):
    dataframe.to_excel(file_path, index=False)

root.state('zoomed')

root.mainloop()

